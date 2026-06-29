# =============================================================================
# Módulo: gestor_excel.py
# Propósito: Abrir, leer y escribir el archivo Excel que contiene la lista
#            de RUC a procesar. Actualiza las columnas de Estado y Detalle.
# Ciclo de vida: Fase 1 (Apertura), Fase 2 (Lectura), Fase 4 (Escritura),
#                Fase 5 (Guardado y cierre)
# =============================================================================

from openpyxl import load_workbook


class GestorExcel:
    """
    Maneja el ciclo de vida del archivo Excel:
      - Abre el libro y la hoja especificada.
      - Itera fila por fila leyendo el RUC.
      - Escribe el resultado (Estado + Detalle) en cada fila procesada.
      - Guarda los cambios y cierra el archivo al finalizar.
    """

    def __init__(self, ruta_excel: str, nombre_hoja: str = "Sheet1",
                 columna_ruc: str = "A", columna_estado: str = "B",
                 columna_detalle: str = "C", columna_nombre: str = "A"):
        """
        Args:
            ruta_excel: Ruta completa al archivo .xlsx
            nombre_hoja: Nombre de la hoja donde están los datos.
            columna_ruc: Letra de la columna donde está el RUC (ej: "B").
            columna_estado: Letra de la columna para escribir Estado (ej: "C").
            columna_detalle: Letra de la columna para escribir Detalle (ej: "D").
            columna_nombre: Letra de la columna con el nombre de empresa (ej: "A").
        """
        self.ruta_excel = ruta_excel
        self.nombre_hoja = nombre_hoja
        self.columna_ruc = columna_ruc
        self.columna_estado = columna_estado
        self.columna_detalle = columna_detalle
        self.columna_nombre = columna_nombre
        self.libro = None
        self.hoja = None
        # El indexador arranca en la fila 2 (asumiendo fila 1 = encabezados)
        self.fila_index = 2
        self.total_filas = 0

    def abrir(self) -> "GestorExcel":
        """
        Carga el libro Excel en memoria usando openpyxl.
        Calcula total_filas basado en la columna de RUC (columna B)
        para asegurar que lee TODAS las filas con datos, ignorando
        si otras columnas (C, D) tienen más o menos filas.

        Retorna self para permitir encadenamiento de métodos.
        """
        self.libro = load_workbook(self.ruta_excel)
        self.hoja = self.libro[self.nombre_hoja]

        # Calcular la última fila con datos en la columna de RUC
        # openpyxl max_row se basa en todas las columnas.
        # Nosotros solo nos interesa hasta donde haya RUC.
        max_ruc = self.hoja.max_row
        for fila in range(self.hoja.max_row, 1, -1):
            celda = self.hoja[f"{self.columna_ruc}{fila}"].value
            if celda is not None and str(celda).strip():
                max_ruc = fila
                break
        # También verificar columna de nombre
        for fila in range(self.hoja.max_row, 1, -1):
            celda = self.hoja[f"{self.columna_nombre}{fila}"].value
            if celda is not None and str(celda).strip():
                if fila > max_ruc:
                    max_ruc = fila
                break

        self.total_filas = max_ruc
        return self

    def obtener_ruc(self) -> str:
        """
        Lee el valor de la columna de RUC en la fila actual.
        La columna se define en config.json (ColumnaRUC).

        Returns:
            str: Número de RUC o cadena vacía si la celda está en blanco.
        """
        return str(self.hoja[f"{self.columna_ruc}{self.fila_index}"].value or "").strip()

    def obtener_nombre_empresa(self) -> str:
        """
        Lee el nombre de la empresa desde la columna A.

        Returns:
            str: Nombre de la empresa limpio o cadena vacía.
        """
        val = self.hoja[f"{self.columna_nombre}{self.fila_index}"].value
        return str(val or "").strip()

    def escribir_estado(self, estado: str, detalle: str):
        """
        Actualiza las columnas de Estado y Detalle en la fila actual.
        Las columnas se definen en config.json (ColumnaEstado, ColumnaDetalle).

        Args:
            estado: "Procesado" o "Error"
            detalle: Descripción del resultado (ej: "Descargado con éxito")
        """
        self.hoja[f"{self.columna_estado}{self.fila_index}"] = estado
        self.hoja[f"{self.columna_detalle}{self.fila_index}"] = detalle

    def avanzar_fila(self):
        """Incrementa el indexador para pasar a la siguiente fila."""
        self.fila_index += 1

    def guardar_y_cerrar(self):
        """
        Persiste todos los cambios realizados en el archivo Excel
        y cierra la conexión con el libro.
        """
        if self.libro:
            self.libro.save(self.ruta_excel)
            self.libro.close()
